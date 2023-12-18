import openpyxl
from random import shuffle

# 단어와 뜻을 저장할 리스트
word_dict = []

# 단어와 뜻 입력 받기
while True:
    word = input("단어를 입력하세요 (종료하려면 '완료'를 입력하세요): ")
    if word == '완료':
        break
    meaning = input(f"{word}의 뜻을 입력하세요: ")
    word_dict.append((word, meaning))

# 시험 문제 생성
word_dict = word_dict[:20]
questions = word_dict.copy()
shuffle(questions)

# 엑셀 워크북 및 워크시트 생성
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = "단어 시험지"

# 문제 출력
for i, (word, meaning) in enumerate(questions, start=1):
    worksheet.cell(row=i, column=1, value=f"{i}. {word}")
    worksheet.cell(row=i, column=2, value=f"답:{meaning}")

# 엑셀 파일 저장
excel_file = "단어_시험지(XX.xx).xlsx"
workbook.save(excel_file)
print(f"단어 시험지 엑셀 파일 '{excel_file}'이 생성되었습니다.")

# *파일 이름을 수정해주시고, 영어 행의 글자를 하얀색으로 설정해주세요